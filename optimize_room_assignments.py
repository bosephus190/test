import pandas as pd
import numpy as np
import os
from datetime import datetime, timedelta

# Settings
OUTPUT_FOLDER = 'Output'
os.makedirs(OUTPUT_FOLDER, exist_ok = True)
TODAY = pd.Timestamp.today().normalize()
OUTPUT_FILENAME = f'Recommended Changes - {TODAY.strftime("%Y.%m.%d")}.xlsx'
OCCUPANCY_THRESHOLD = 0.7 # easy to adjust
SPREAD_BUFFER = 0.90 # spread revenue buffer → rooms fill to 90% of target

# Home type upgrade priority
home_type_priority = ['B1QL', 'G4K', 'G4KP', 'G5K', 'G5KH', 'G5KO', 'G5KP', 'W5KO', 'W6KO', 'V6K', 'V6KR']

# Room Type → Home Type mapping(normalized)

UPGRADE_PATH = {
'B1QL': ['G4K', 'G4KP'],
'G4K': ['G4KP', 'G5K'],
'G4KP': ['G5K'],
'G5K': ['G5KH', 'G5KO', 'G5KP'],
'G5KH': ['G5KO', 'G5KP', 'W5KO'],
'G5KO': ['G5KP', 'W5KO'],
'G5KP': ['W5KO'],
'W5KO': ['W6KO'],
'W6KO': ['V6K'],
'V6K': ['V6KR']
}

ROOM_TO_HOME_TYPE = {
'B1QL': 'Surf Bungalow',
'G4K': 'Resort Cottage', 'G4KP': 'Resort Cottage', 'G5K': 'Resort Cottage',
'G5KH': 'Resort Cottage', 'G5KO': 'Resort Cottage', 'G5KP': 'Resort Cottage',
'V6K': 'Beach Villa', 'V6KR': 'Beach Villa',
'W5KO': 'Waterside Retreat', 'W6KO': 'Waterside Retreat'
}

home_type_rank = {ht: i for i, ht in enumerate(home_type_priority)}
get_stay_dates = lambda arrival, departure: [d.date()for d in pd.date_range(arrival, departure - timedelta(days = 1))]

# Load data
sheet1_df = pd.read_excel('Sheet1.xlsx')
availability_df = pd.read_excel('Availability.xlsx')

# Clean stays
df = sheet1_df[~sheet1_df['Status'].isin(['Cancelled', 'No Show'])].copy()
df['Arrival Date'] = pd.to_datetime(df['Arrival Date']).dt.normalize()
df['Departure Date'] = pd.to_datetime(df['Departure Date']).dt.normalize()
df['Room Type'] = df['Room Type'].astype(str)
df['Room Number'] = df['Room'].astype(str)
df['Home Type'] = df['Room Type'].map(ROOM_TO_HOME_TYPE)

# = = = Explode Stays to Nightly Rate for Year-Based Targeting = = =
expanded_rows = []
for _, row in df.iterrows():
 stay_nights =(row['Departure Date'] - row['Arrival Date']).days
 if stay_nights = = 0:
 continue
 nightly_rate = row['Total Room Rate'] / stay_nights
 for single_date in pd.date_range(row['Arrival Date'], row['Departure Date'] - timedelta(days = 1)):
 expanded_rows.append({
 'Room Type': row['Room Type'],
 'Date': single_date.date(),
 'Total Rate': nightly_rate
 })

 expanded_df = pd.DataFrame(expanded_rows)
 expanded_df['Year'] = pd.to_datetime(expanded_df['Date']).dt.year
 expanded_df['Home Type'] = expanded_df['Room Type'].map(ROOM_TO_HOME_TYPE)

 # Process availability
 availability_df = availability_df[availability_df['Availability'] = = 'Available'].copy()
 availability_df['First Date of Availability'] = pd.to_datetime(availability_df['First Date of Availability']).dt.normalize()
 availability_df['Comp'] = availability_df['Comp'].fillna('').str.upper().eq('YES')

 # Build room availability dict
 room_info = {}
 for _, row in availability_df.iterrows():
 room_number = str(row['Homes'])
 room_type = row['Room Type']
 first_avail = row['First Date of Availability']
 is_comp = row['Comp']
 room_info[room_number] = {
 'Room Type': room_type,
 'Home Type': ROOM_TO_HOME_TYPE.get(room_type),
 'First Available': first_avail,
 'Comp': is_comp
 }

 # Build initial room calendar
 room_calendar = {room: set()for room in room_info}

 for _, row in df.iterrows():
 room = row['Room Number']
 if pd.notnull(room)and room ! = 'nan' and room in room_calendar:
 for single_date in pd.date_range(row['Arrival Date'], row['Departure Date'] - timedelta(days = 1)):
 room_calendar[room].add(single_date.date())

 # Owner and Comp Rate Plans
 owner_rate_plans = {'4RCOWN', 'BUNOWN', 'BVOWN', 'RCOWN', 'WROWN'}
 comp_rate_plan = 'COMP'

 # Build home type revenue per year
 df['Year'] = pd.to_datetime(df['Arrival Date']).dt.year
 home_type_year_revenue = expanded_df.groupby(['Room Type', 'Year'])['Total Rate'].sum().reset_index()

 # Build target revenue per room per year
 target_revenue = []
 for room, info in room_info.items():
 room_type = info['Room Type']
 first_avail = info['First Available']

 for year in df['Year'].unique():
 total_rev_row = home_type_year_revenue[(home_type_year_revenue['Room Type'] = = room_type)&(home_type_year_revenue['Year'] = = year)]
 if total_rev_row.empty:
 continue
 total_rev = total_rev_row['Total Rate'].values[0]

 days_in_year = 366 if year % 4 = = 0 else 365

 if first_avail.year > year:
 avail_fraction = 0.0
 elif first_avail.year = = year and first_avail > pd.Timestamp(f"{year}-01-01"):
 avail_fraction =(days_in_year -(first_avail - pd.Timestamp(f"{year}-01-01")).days)/ days_in_year
 else:
 avail_fraction = 1.0

 available_rooms_count = len([r for r, i in room_info.items()if i['Room Type'] = = room_type and(i['First Available'].year < year or(i['First Available'].year = = year and i['First Available'] < = pd.Timestamp(f"{year}-01-01")))])

 if available_rooms_count = = 0:
 target_rev = 0.0
 else:
 target_rev = avail_fraction * total_rev / available_rooms_count

 target_revenue.append({
 'Room Number': room,
 'Year': year,
 'Target Revenue': target_rev
 })

 target_revenue_df = pd.DataFrame(target_revenue)

 # Initialize current revenue tracker
 current_revenue = {(row['Room Number'], row['Year']): 0.0 for row in target_revenue}
 # Helper: calculate daily occupancy(unchanged)
 def calculate_daily_occupancy(df, assigned_rooms):
 occupancy = {}
 for _, row in df.iterrows():
 home_type = row['Room Type']
 room = assigned_rooms.get(row['Confirmation Number'], row['Room Number'])
 if pd.isna(room)or room = = 'nan':
 continue

 rate_plan = row['Rate Plan']
 for single_date in pd.date_range(row['Arrival Date'], row['Departure Date'] - timedelta(days = 1)):
 date = single_date.date()
 key =(home_type, date)
 if key not in occupancy:
 occupancy[key] = {'Owner': 0, 'Comp': 0, 'Transient': 0}

 if rate_plan in owner_rate_plans:
 occupancy[key]['Owner'] + = 1
 elif rate_plan = = comp_rate_plan:
 occupancy[key]['Comp'] + = 1
 else:
 occupancy[key]['Transient'] + = 1
 return occupancy

 def find_available_room(home_type, arrival, departure, is_comp, skip_rooms, reason, year):
 if reason = = 'Spread Revenue':
 # Only use same home type during Spread Revenue
 possible_rooms = [r for r, info in room_info.items()
 if info['Room Type'] = = home_type
 and(not is_comp or info['Comp'])
 and info['First Available'] < = arrival
 and r not in skip_rooms]

 room_target_pairs = []
 for room in possible_rooms:
 target_row = target_revenue_df[
(target_revenue_df['Room Number'] = = room)&(target_revenue_df['Year'] = = year)]
 if target_row.empty:
 continue
 target_rev = target_row['Target Revenue'].values[0]
 curr_rev = current_revenue.get((room, year), 0.0)

 if target_rev = = 0:
 continue

 pct_filled = curr_rev / target_rev
 if pct_filled > = SPREAD_BUFFER:
 continue # ✅ New: Don't allow rooms above the spread buffer

 room_target_pairs.append((room, pct_filled, curr_rev))

 # ✅ Sort by % of Target filled(asc), then by raw current revenue(asc)
 room_target_pairs.sort(key = lambda x: (x[1], x[2]))

 for room, pct_filled, _ in room_target_pairs:
 room_dates = room_calendar[room]
 conflict = any(date in room_dates for date in get_stay_dates(arrival, departure))
 if not conflict:
 return room

 else:
 # Consolidate phase → allow upgrades
 upgradeable_types = [home_type] + UPGRADE_PATH.get(home_type, [])
 possible_rooms = [r for r, info in room_info.items()
 if info['Room Type'] in upgradeable_types
 and(not is_comp or info['Comp'])
 and info['First Available'] < = arrival
 and r not in skip_rooms]

 possible_rooms = sorted(possible_rooms, key = lambda r: home_type_rank[room_info[r]['Room Type']])
 for room in possible_rooms:
 room_dates = room_calendar[room]
 conflict = any(date in room_dates for date in get_stay_dates(arrival, departure))
 if not conflict:
 return room

 return None
 # Phase 1 → Consolidate loop(≥ 70%)
 print("\n = = = Phase 1 → Consolidate(≥ 70%)= = = ")
 df_sorted = df.copy()
 df_sorted['Nights'] =(pd.to_datetime(df_sorted['Departure Date'])- pd.to_datetime(df_sorted['Arrival Date'])).dt.days.replace(0, 1)
 df_sorted['SortKey'] = df_sorted['Total Room Rate'] / df_sorted['Nights']
 df_sorted = df_sorted.sort_values(by = 'SortKey', ascending = False)

 assigned_rooms = {}
 change_recommendations = []
 cascade_steps = {}
 occupancy = calculate_daily_occupancy(df, assigned_rooms)

 pass_number = 0
 changes_made = True

 while changes_made:
 print(f"\n--- Phase 1 → Pass {pass_number} ---")
 changes_made = False

 for _, row in df_sorted.iterrows():
 conf = row['Confirmation Number']
 arrival = row['Arrival Date']
 departure = row['Departure Date']
 room = row['Room Number']
 home_type = row['Room Type']
 do_not_move = str(row.get('Do Not Move', '')).upper()= = 'YES'
 is_comp = row['Rate Plan'] = = 'COMP'
 year = row['Year']

 if arrival < TODAY:
 continue

 current_room = assigned_rooms.get(conf, room)

 available_rooms = [r for r in room_info if room_info[r]['Room Type'] = = home_type]

 over_occupancy = any(
(occupancy.get((home_type, d), {'Owner': 0, 'Comp': 0, 'Transient': 0})['Transient'] /
 max(1, len(available_rooms)-
 occupancy.get((home_type, d), {'Owner': 0, 'Comp': 0, 'Transient': 0})['Owner'] -
 occupancy.get((home_type, d), {'Owner': 0, 'Comp': 0, 'Transient': 0})['Comp']))> OCCUPANCY_THRESHOLD
 for d in pd.date_range(arrival, departure - timedelta(days = 1)).date)

 reason = None
 target_room = None

 if(pd.isna(room)or room = = 'nan')and assigned_rooms.get(conf)is None:
 reason = 'Initial Room Assignment'
 elif do_not_move:
 assigned_rooms[conf] = current_room
 continue
 elif over_occupancy:
 reason = 'Consolidate'
 else:
 continue # Skip → Phase 1 only moves for Consolidate

 skip_rooms = set(r for r, info in room_info.items()if info['First Available'] > arrival)
 target_room = find_available_room(home_type, arrival, departure, is_comp, skip_rooms, reason, year)

 cascade_step = pass_number

 if target_room and target_room ! = current_room:
 for single_date in pd.date_range(arrival, departure - timedelta(days = 1)):
 room_calendar[target_room].add(single_date.date())
 assigned_rooms[conf] = target_room

 print(f"Moving CONF {conf}({reason})to room {target_room}(Cascade Step {cascade_step})")

 change_recommendations.append({
 'Confirmation Number': conf,
 'Arrival Date': arrival,
 'Departure Date': departure,
 'Home Type': row['Home Type'],
 'Original Room Number': room,
 'Recommended Room Number': target_room,
 'Total Rate': row['Total Room Rate'],
 'Reason for Change': reason,
 'Cascade Step': cascade_step
 })

 current_revenue[(target_room, year)] + = row['Total Room Rate']
 changes_made = True

 occupancy = calculate_daily_occupancy(df, assigned_rooms)
 pass_number + = 1

 print("\n = = = Phase 1 complete = = = \n")
 def recompute_dynamic_target_revenue(df, assigned_rooms, room_info):
 # First → compute Total Current Revenue per home type
 current_home_type_revenue = {}
 for _, row in df.iterrows():
 # Skip if room type occupancy is already > 70%
 for room in assigned_rooms.values():
 if room not in room_info:
 continue
 room_type = room_info[room]['Room Type']
 type_rooms = [r for r in room_info if room_info[r]['Room Type'] = = room_type]
 if room not in room_info:
 continue # Skip if room info is missing
 room_type = room_info[room]['Room Type']
 type_rooms = [r for r in room_info if room_info[r]['Room Type'] = = room_type]
 if room not in room_info:
 continue # Skip invalid or untracked rooms
 if room not in room_info:
 continue # Skip invalid or untracked rooms
 occupied_nights = sum(len(room_calendar.get(r, set()))for r in type_rooms)
 total_nights = len(type_rooms)* 365
 occupancy = occupied_nights / total_nights
 if occupancy > 0.7:
 continue
 conf = row['Confirmation Number']
 final_room = assigned_rooms.get(conf, row['Room Number'])
 home_type_final = room_info[final_room]['Room Type']
 year = row['Year']

 key =(home_type_final, year)
 current_home_type_revenue[key] = current_home_type_revenue.get(key, 0.0)+ row['Total Room Rate']

 # Second → compute Available Fraction per room(unchanged from original logic)
 target_revenue_list = []
 for room, info in room_info.items():
 room_type = info['Room Type']
 first_avail = info['First Available']

 for year in df['Year'].unique():
 days_in_year = 366 if year % 4 = = 0 else 365

 if first_avail.year > year:
 avail_fraction = 0.0
 elif first_avail.year = = year and first_avail > pd.Timestamp(f"{year}-01-01"):
 avail_fraction =(days_in_year -(first_avail - pd.Timestamp(f"{year}-01-01")).days)/ days_in_year
 else:
 avail_fraction = 1.0

 # Now compute Total Available Fraction for current rooms in this home type and year
 available_rooms = [
 r for r, i in room_info.items()
 if i['Room Type'] = = room_type
 and(i['First Available'].year < year or(i['First Available'].year = = year and i['First Available'] < = pd.Timestamp(f"{year}-01-01")))
 ]

 total_available_fraction = 0.0
 for r in available_rooms:
 first_avail_r = room_info[r]['First Available']
 if first_avail_r.year > year:
 f = 0.0
 elif first_avail_r.year = = year and first_avail_r > pd.Timestamp(f"{year}-01-01"):
 f =(days_in_year -(first_avail_r - pd.Timestamp(f"{year}-01-01")).days)/ days_in_year
 else:
 f = 1.0
 total_available_fraction + = f

 # Now compute Target Revenue for this room
 key =(room_type, year)
 total_current_revenue = current_home_type_revenue.get(key, 0.0)

 if total_available_fraction = = 0:
 target_rev = 0.0
 else:
 target_rev = avail_fraction * total_current_revenue / total_available_fraction

 target_revenue_list.append({
 'Room Number': room,
 'Year': year,
 'Target Revenue': target_rev
 })

 # Return as DataFrame → ready to use in Spread Revenue loop
 return pd.DataFrame(target_revenue_list)

 # = = = = = Final Phase 2 Loop: Spread Revenue with all improvements = = = = =

 print("\n = = = Phase 2 → Spread Revenue(< 70%)= = = ")
 df_sorted = df.copy()
 df_sorted['Nights'] =(pd.to_datetime(df_sorted['Departure Date'])- pd.to_datetime(df_sorted['Arrival Date'])).dt.days.replace(0, 1)
 df_sorted['SortKey'] = df_sorted['Total Room Rate'] / df_sorted['Nights']
 df_sorted = df_sorted.sort_values(by = 'SortKey', ascending = False)

 pass_number = 0
 changes_made = True

 while changes_made:
 print(f"\n--- Phase 2 → Pass {pass_number} ---")
 changes_made = False

 for _, row in df_sorted.iterrows():
 conf = row['Confirmation Number']
 arrival = row['Arrival Date']
 departure = row['Departure Date']
 room = row['Room Number']
 home_type = row['Room Type']
 do_not_move = str(row.get('Do Not Move', '')).upper()= = 'YES'
 is_comp = row['Rate Plan'] = = 'COMP'
 year = row['Year']

 if arrival < TODAY:
 continue

 current_room = assigned_rooms.get(conf, room)

 available_rooms = [r for r in room_info if room_info[r]['Room Type'] = = home_type]

 # Improved → Mean occupancy % instead of strict all()check
 occupancy_pct_list = [
(occupancy.get((home_type, d), {'Owner': 0, 'Comp': 0, 'Transient': 0})['Transient'] /
 max(1, len(available_rooms)-
 occupancy.get((home_type, d), {'Owner': 0, 'Comp': 0, 'Transient': 0})['Owner'] -
 occupancy.get((home_type, d), {'Owner': 0, 'Comp': 0, 'Transient': 0})['Comp']))
 for d in pd.date_range(arrival, departure - timedelta(days = 1)).date
 ]

 occupancy_below_threshold =(np.mean(occupancy_pct_list)< OCCUPANCY_THRESHOLD)

 reason = None
 target_room = None

 if(pd.isna(room)or room = = 'nan')and assigned_rooms.get(conf)is None:
 reason = 'Initial Room Assignment'
 elif do_not_move:
 assigned_rooms[conf] = current_room
 continue
 elif occupancy_below_threshold:
 reason = 'Spread Revenue'
 else:
 continue

 skip_rooms = set(r for r, info in room_info.items()if info['First Available'] > arrival)
 target_room = find_available_room(home_type, arrival, departure, is_comp, skip_rooms, reason, year)

 cascade_step = pass_number

 if target_room and target_room ! = current_room:
 for single_date in pd.date_range(arrival, departure - timedelta(days = 1)):
 room_calendar[target_room].add(single_date.date())
 assigned_rooms[conf] = target_room

 print(f"Moving CONF {conf}({reason})to room {target_room}(Cascade Step {cascade_step})")

 change_recommendations.append({
 'Confirmation Number': conf,
 'Arrival Date': arrival,
 'Departure Date': departure,
 'Home Type': row['Home Type'],
 'Original Room Number': room,
 'Recommended Room Number': target_room,
 'Total Rate': row['Total Room Rate'],
 'Reason for Change': reason,
 'Cascade Step': cascade_step
 })

 current_revenue[(target_room, year)] + = row['Total Room Rate']
 changes_made = True

 occupancy = calculate_daily_occupancy(df, assigned_rooms)

 # 🚀 Dynamic Target Revenue recompute after each pass
 target_revenue_df = recompute_dynamic_target_revenue(df, assigned_rooms, room_info)

 pass_number + = 1

 # Optional → shuffle stays each pass → avoids early stay bias
 df_sorted = df_sorted.sample(frac = 1, random_state = pass_number).reset_index(drop = True)

 print("\n = = = Phase 2 complete = = = \n")

 target_revenue_df = recompute_dynamic_target_revenue(df, assigned_rooms, room_info)

 print("\n = = = Phase 3 → Rebalance Overfilled Rooms to Underfilled Ones = = = ")

 # Recompute Target Revenue and Occupancy again(safe refresh)
 target_revenue_df = recompute_dynamic_target_revenue(df, assigned_rooms, room_info)
 occupancy = calculate_daily_occupancy(df, assigned_rooms)

 # Build Target and Current Revenue lookup
 target_dict = {(row['Room Number'], row['Year']): row['Target Revenue'] for _, row in target_revenue_df.iterrows()}

 # Classify overfilled and underfilled rooms
 overfilled = {}
 underfilled = {}
 for(room, year), target in target_dict.items():
 curr = sum(
 row['Total Room Rate'] for _, row in df.iterrows()
 if assigned_rooms.get(row['Confirmation Number'], row['Room Number'])= = room and row['Year'] = = year
)
 if target = = 0:
 continue
 pct = curr / target
 if pct > 1.10:
 overfilled[(room, year)] = pct
 elif pct < 0.90:
 underfilled[(room, year)] = pct

 rebalance_moves = 0

 # For each overfilled room, try to move one stay to underfilled rooms
 for(over_room, year), over_pct in sorted(overfilled.items(), key = lambda x: -x[1]):
 over_home_type = room_info[over_room]['Room Type']

 for(over_room, year), over_pct in sorted(overfilled.items(), key = lambda x: -x[1]):
 over_home_type = room_info[over_room]['Room Type']
 pass # placeholder to ensure valid block
 occupied_nights = sum(len(room_calendar.get(r, set()))for r in type_rooms)

available_nights = sum(
(datetime(year + 1, 1, 1).date()- max(first_avail.get(r, datetime(year, 1, 1).date()), datetime(year, 1, 1).date())).days
 for r in type_rooms
)
if available_nights = = 0 or occupied_nights / available_nights > 0.70:
 continue
 continue
 if row['Rate Plan'] in owner_rate_plans or row['Rate Plan'] = = comp_rate_plan:
 if available_nights = = 0:
 continue

 arrival = row['Arrival Date']
 departure = row['Departure Date']
 is_comp = row['Rate Plan'] = = 'COMP'
 moved = False

 for(under_room, u_pct)in sorted(
 for under_room, u_pct in sorted(
 [(r, p)for(r, y), p in underfilled.items()],
 key = lambda x: x[1]
):
 # placeholder for rebalancing logic
 continue
 continue

 # Check for date conflicts
 room_dates = room_calendar[under_room]
 conflict = any(date in room_dates for date in get_stay_dates(arrival, departure))
 if not conflict:
 # ✅ Perform rebalancing move
 assigned_rooms[conf] = under_room
 for d in pd.date_range(arrival, departure - timedelta(days = 1)): )
 room_calendar[under_room].add(d.date())
 rebalance_moves + = 1
 moved = True

 change_recommendations.append({
 'Confirmation Number': conf,
 'Arrival Date': arrival,
 'Departure Date': departure,
 'Home Type': row['Home Type'],
 'Original Room Number': over_room,
 'Recommended Room Number': under_room,
 'Total Rate': row['Total Room Rate'],
 'Reason for Change': 'Rebalance',
 'Cascade Step': 'Phase 3'
 })

 print(f"→ Rebalanced CONF {conf} from {over_room}({over_pct: .1%})→ {under_room}({u_pct: .1%})")
 # continue trying other stays in the same room to reduce overfill

 if not moved:
 print(f"✘ Could not rebalance CONF {conf} from {over_room} — no eligible underfilled room found.")

 print(f"\n = = = Phase 3 complete — {rebalance_moves} rebalancing moves made = = = \n")

 target_revenue_df = recompute_dynamic_target_revenue(df, assigned_rooms, room_info)

 # Fix: Convert Cascade Step to string to avoid type sorting error
 df_changes = pd.DataFrame(change_recommendations)
 if __name__ = = "__main__":
 df_changes['Cascade Step'] = df_changes['Cascade Step'].astype(str)

 # Keep only the last recommendation per confirmation number
 df_changes_final = df_changes.sort_values(by = 'Cascade Step').groupby('Confirmation Number', as_index = False).last()

 # Recompute Current Revenue → Final Revenue per room → based on final room assignments
 print("\n = = = Recomputing Final Revenue per Room = = = ")

 current_revenue = {(row['Room Number'], row['Year']): 0.0 for _, row in target_revenue_df.iterrows()}

 for _, row in df.iterrows():
 # Skip if room type occupancy is already > 70%
 pass # placeholder block for loop
 total_nights = len(type_rooms)* 365
 occupancy = occupied_nights / total_nights
 if occupancy > 0.7:
 if available_nights = = 0:
 continue
 arrival = row['Arrival Date']
 departure = row['Departure Date']
 final_room = assigned_rooms.get(conf, row['Room Number'])
 year = row['Year']

 # Add full stay revenue to final_room's Current Revenue
 current_revenue[(final_room, year)] + = row['Total Room Rate']

 # Compute room revenue & occupancy
 room_stats = []
 for(room, year), target in target_dict.items():
 actual = current_revenue.get((room, year), 0.0)
 occupancy_nights = sum(
(row['Departure Date'] - row['Arrival Date']).days
 for _, row in df.iterrows()
 if assigned_rooms.get(row['Confirmation Number'], row['Room Number'])= = room and row['Year'] = = year
 and row['Rate Plan'] not in owner_rate_plans | {comp_rate_plan}
)

 available_nights = 365
 first_avail = room_info[room]['First Available']
 if first_avail.year = = year:
 available_nights =(pd.Timestamp(f"{year + 1}-01-01")- max(first_avail, pd.Timestamp(f"{year}-01-01"))).days
 elif first_avail.year > year:
 available_nights = 0

 original_occupancy_pct =(occupancy_nights / available_nights)if available_nights > 0 else 0
 actual_pct_of_target = actual / target if target > 0 else 0

 print(f"Room {room} | Year {year} | Target: ${target: , .2f} | Actual: ${actual: , .2f} | {actual_pct_of_target: .1%} of Target")

 room_stats.append({
 'Year': year,
 'Room Number': room,
 'Target Revenue': target,
 'Final Room Revenue': actual,
 'Final % of Target': actual_pct_of_target,
 'Final Occupancy %': original_occupancy_pct
 })

 room_stats_df = pd.DataFrame(room_stats)

 print("\n = = = Validating for overlapping stays(Auto-Fix Enabled)= = = ")
 conflict_count = 0
 fix_count = 0
 unfixable = 0
 room_night_tracker = {}
 unfixable_conflicts = []

 for _, row in df.iterrows():
 # Skip if room type occupancy is already > 70%
 occupied_nights = sum(len(room_calendar.get(r, set()))for r in type_rooms)
 total_nights = len(type_rooms)* 365
 occupancy = occupied_nights / total_nights
 if occupancy > 0.7:
 continue # completes `if occupancy > 0.7: `
 if available_nights = = 0:
 continue
 continue
 arrival = row['Arrival Date']
 departure = row['Departure Date']
 assigned_room = assigned_rooms.get(conf, row['Room Number'])
 home_type = row['Room Type']
 is_comp = row['Rate Plan'] = = comp_rate_plan
 year = row['Year']

 import pandas as pd
 import numpy as np
 import os
 from datetime import datetime, timedelta

 # Settings
 OUTPUT_FOLDER = 'Output'
 os.makedirs(OUTPUT_FOLDER, exist_ok = True)
 TODAY = pd.Timestamp.today().normalize()
 OUTPUT_FILENAME = f'Recommended Changes - {TODAY.strftime("%Y.%m.%d")}.xlsx'
 OCCUPANCY_THRESHOLD = 0.7 # easy to adjust
 SPREAD_BUFFER = 0.90 # spread revenue buffer → rooms fill to 90% of target

 # Home type upgrade priority
 home_type_priority = ['B1QL', 'G4K', 'G4KP', 'G5K', 'G5KH', 'G5KO', 'G5KP', 'W5KO', 'W6KO', 'V6K', 'V6KR']

 # Room Type → Home Type mapping(normalized)

 UPGRADE_PATH = {
 'B1QL': ['G4K', 'G4KP'],
 'G4K': ['G4KP', 'G5K'],
 'G4KP': ['G5K'],
 'G5K': ['G5KH', 'G5KO', 'G5KP'],
 'G5KH': ['G5KO', 'G5KP', 'W5KO'],
 'G5KO': ['G5KP', 'W5KO'],
 'G5KP': ['W5KO'],
 'W5KO': ['W6KO'],
 'W6KO': ['V6K'],
 'V6K': ['V6KR']
 }

 ROOM_TO_HOME_TYPE = {
 'B1QL': 'Surf Bungalow',
 'G4K': 'Resort Cottage', 'G4KP': 'Resort Cottage', 'G5K': 'Resort Cottage',
 'G5KH': 'Resort Cottage', 'G5KO': 'Resort Cottage', 'G5KP': 'Resort Cottage',
 'V6K': 'Beach Villa', 'V6KR': 'Beach Villa',
 'W5KO': 'Waterside Retreat', 'W6KO': 'Waterside Retreat'
 }

 home_type_rank = {ht: i for i, ht in enumerate(home_type_priority)}
 get_stay_dates = lambda arrival, departure: [d.date()for d in pd.date_range(arrival, departure - timedelta(days = 1))]

 # Load data
 sheet1_df = pd.read_excel('Sheet1.xlsx')
 availability_df = pd.read_excel('Availability.xlsx')

 # Clean stays
 df = sheet1_df[~sheet1_df['Status'].isin(['Cancelled', 'No Show'])].copy()
 df['Arrival Date'] = pd.to_datetime(df['Arrival Date']).dt.normalize()
 df['Departure Date'] = pd.to_datetime(df['Departure Date']).dt.normalize()
 df['Room Type'] = df['Room Type'].astype(str)
 df['Room Number'] = df['Room'].astype(str)
 df['Home Type'] = df['Room Type'].map(ROOM_TO_HOME_TYPE)

 # = = = Explode Stays to Nightly Rate for Year-Based Targeting = = =
 expanded_rows = []
 for _, row in df.iterrows():
 stay_nights =(row['Departure Date'] - row['Arrival Date']).days
 if stay_nights = = 0:
 continue
 nightly_rate = row['Total Room Rate'] / stay_nights
 for single_date in pd.date_range(row['Arrival Date'], row['Departure Date'] - timedelta(days = 1)):
 expanded_rows.append({
 'Room Type': row['Room Type'],
 'Date': single_date.date(),
 'Total Rate': nightly_rate
 })

 expanded_df = pd.DataFrame(expanded_rows)
 expanded_df['Year'] = pd.to_datetime(expanded_df['Date']).dt.year
 expanded_df['Home Type'] = expanded_df['Room Type'].map(ROOM_TO_HOME_TYPE)

 # Process availability
 availability_df = availability_df[availability_df['Availability'] = = 'Available'].copy()
 availability_df['First Date of Availability'] = pd.to_datetime(availability_df['First Date of Availability']).dt.normalize()
 availability_df['Comp'] = availability_df['Comp'].fillna('').str.upper().eq('YES')

 # Build room availability dict
 room_info = {}
 for _, row in availability_df.iterrows():
 room_number = str(row['Homes'])
 room_type = row['Room Type']
 first_avail = row['First Date of Availability']
 is_comp = row['Comp']
 room_info[room_number] = {
 'Room Type': room_type,
 'Home Type': ROOM_TO_HOME_TYPE.get(room_type),
 'First Available': first_avail,
 'Comp': is_comp
 }

 # Build initial room calendar
 room_calendar = {room: set()for room in room_info}

 for _, row in df.iterrows():
 room = row['Room Number']
 if pd.notnull(room)and room ! = 'nan' and room in room_calendar:
 for single_date in pd.date_range(row['Arrival Date'], row['Departure Date'] - timedelta(days = 1)):
 room_calendar[room].add(single_date.date())

 # Owner and Comp Rate Plans
 owner_rate_plans = {'4RCOWN', 'BUNOWN', 'BVOWN', 'RCOWN', 'WROWN'}
 comp_rate_plan = 'COMP'

 # Build home type revenue per year
 df['Year'] = pd.to_datetime(df['Arrival Date']).dt.year
 home_type_year_revenue = expanded_df.groupby(['Room Type', 'Year'])['Total Rate'].sum().reset_index()

 # Build target revenue per room per year
 target_revenue = []
 for room, info in room_info.items():
 room_type = info['Room Type']
 first_avail = info['First Available']

 for year in df['Year'].unique():
 total_rev_row = home_type_year_revenue[(home_type_year_revenue['Room Type'] = = room_type)&(home_type_year_revenue['Year'] = = year)]
 if total_rev_row.empty:
 continue
 total_rev = total_rev_row['Total Rate'].values[0]

 days_in_year = 366 if year % 4 = = 0 else 365

 if first_avail.year > year:
 avail_fraction = 0.0
 elif first_avail.year = = year and first_avail > pd.Timestamp(f"{year}-01-01"):
 avail_fraction =(days_in_year -(first_avail - pd.Timestamp(f"{year}-01-01")).days)/ days_in_year
 else:
 avail_fraction = 1.0

 available_rooms_count = len([r for r, i in room_info.items()if i['Room Type'] = = room_type and(i['First Available'].year < year or(i['First Available'].year = = year and i['First Available'] < = pd.Timestamp(f"{year}-01-01")))])

 if available_rooms_count = = 0:
 target_rev = 0.0
 else:
 target_rev = avail_fraction * total_rev / available_rooms_count

 target_revenue.append({
 'Room Number': room,
 'Year': year,
 'Target Revenue': target_rev
 })

 target_revenue_df = pd.DataFrame(target_revenue)

 # Initialize current revenue tracker
 current_revenue = {(row['Room Number'], row['Year']): 0.0 for row in target_revenue}
 # Helper: calculate daily occupancy(unchanged)
 def calculate_daily_occupancy(df, assigned_rooms):
 occupancy = {}
 for _, row in df.iterrows():
 home_type = row['Room Type']
 room = assigned_rooms.get(row['Confirmation Number'], row['Room Number'])
 if pd.isna(room)or room = = 'nan':
 continue

 rate_plan = row['Rate Plan']
 for single_date in pd.date_range(row['Arrival Date'], row['Departure Date'] - timedelta(days = 1)):
 date = single_date.date()
 key =(home_type, date)
 if key not in occupancy:
 occupancy[key] = {'Owner': 0, 'Comp': 0, 'Transient': 0}

 if rate_plan in owner_rate_plans:
 occupancy[key]['Owner'] + = 1
 elif rate_plan = = comp_rate_plan:
 occupancy[key]['Comp'] + = 1
 else:
 occupancy[key]['Transient'] + = 1
 return occupancy

 def find_available_room(home_type, arrival, departure, is_comp, skip_rooms, reason, year):
 if reason = = 'Spread Revenue':
 # Only use same home type during Spread Revenue
 possible_rooms = [r for r, info in room_info.items()
 if info['Room Type'] = = home_type
 and(not is_comp or info['Comp'])
 and info['First Available'] < = arrival
 and r not in skip_rooms]

 room_target_pairs = []
 for room in possible_rooms:
 target_row = target_revenue_df[
(target_revenue_df['Room Number'] = = room)&(target_revenue_df['Year'] = = year)]
 if target_row.empty:
 continue
 target_rev = target_row['Target Revenue'].values[0]
 curr_rev = current_revenue.get((room, year), 0.0)

 if target_rev = = 0:
 continue

 pct_filled = curr_rev / target_rev
 if pct_filled > = SPREAD_BUFFER:
 continue # ✅ New: Don't allow rooms above the spread buffer

 room_target_pairs.append((room, pct_filled, curr_rev))

 # ✅ Sort by % of Target filled(asc), then by raw current revenue(asc)
 room_target_pairs.sort(key = lambda x: (x[1], x[2]))

 for room, pct_filled, _ in room_target_pairs:
 room_dates = room_calendar[room]
 conflict = any(date in room_dates for date in get_stay_dates(arrival, departure))
 if not conflict:
 return room

 else:
 # Consolidate phase → allow upgrades
 upgradeable_types = [home_type] + UPGRADE_PATH.get(home_type, [])
 possible_rooms = [r for r, info in room_info.items()
 if info['Room Type'] in upgradeable_types
 and(not is_comp or info['Comp'])
 and info['First Available'] < = arrival
 and r not in skip_rooms]

 possible_rooms = sorted(possible_rooms, key = lambda r: home_type_rank[room_info[r]['Room Type']])
 for room in possible_rooms:
 room_dates = room_calendar[room]
 conflict = any(date in room_dates for date in get_stay_dates(arrival, departure))
 if not conflict:
 return room

 return None
 # Phase 1 → Consolidate loop(≥ 70%)
 print("\n = = = Phase 1 → Consolidate(≥ 70%)= = = ")
 df_sorted = df.copy()
 df_sorted['Nights'] =(pd.to_datetime(df_sorted['Departure Date'])- pd.to_datetime(df_sorted['Arrival Date'])).dt.days.replace(0, 1)
 df_sorted['SortKey'] = df_sorted['Total Room Rate'] / df_sorted['Nights']
 df_sorted = df_sorted.sort_values(by = 'SortKey', ascending = False)

 assigned_rooms = {}
 change_recommendations = []
 cascade_steps = {}
 occupancy = calculate_daily_occupancy(df, assigned_rooms)

 pass_number = 0
 changes_made = True

 while changes_made:
 print(f"\n--- Phase 1 → Pass {pass_number} ---")
 changes_made = False

 for _, row in df_sorted.iterrows():
 conf = row['Confirmation Number']
 arrival = row['Arrival Date']
 departure = row['Departure Date']
 room = row['Room Number']
 home_type = row['Room Type']
 do_not_move = str(row.get('Do Not Move', '')).upper()= = 'YES'
 is_comp = row['Rate Plan'] = = 'COMP'
 year = row['Year']

 if arrival < TODAY:
 continue

 current_room = assigned_rooms.get(conf, room)

 available_rooms = [r for r in room_info if room_info[r]['Room Type'] = = home_type]

 over_occupancy = any(
(occupancy.get((home_type, d), {'Owner': 0, 'Comp': 0, 'Transient': 0})['Transient'] /
 max(1, len(available_rooms)-
 occupancy.get((home_type, d), {'Owner': 0, 'Comp': 0, 'Transient': 0})['Owner'] -
 occupancy.get((home_type, d), {'Owner': 0, 'Comp': 0, 'Transient': 0})['Comp']))> OCCUPANCY_THRESHOLD
 for d in pd.date_range(arrival, departure - timedelta(days = 1)).date)

 reason = None
 target_room = None

 if(pd.isna(room)or room = = 'nan')and assigned_rooms.get(conf)is None:
 reason = 'Initial Room Assignment'
 elif do_not_move:
 assigned_rooms[conf] = current_room
 continue
 elif over_occupancy:
 reason = 'Consolidate'
 else:
 continue # Skip → Phase 1 only moves for Consolidate

 skip_rooms = set(r for r, info in room_info.items()if info['First Available'] > arrival)
 target_room = find_available_room(home_type, arrival, departure, is_comp, skip_rooms, reason, year)

 cascade_step = pass_number

 if target_room and target_room ! = current_room:
 for single_date in pd.date_range(arrival, departure - timedelta(days = 1)):
 room_calendar[target_room].add(single_date.date())
 assigned_rooms[conf] = target_room

 print(f"Moving CONF {conf}({reason})to room {target_room}(Cascade Step {cascade_step})")

 change_recommendations.append({
 'Confirmation Number': conf,
 'Arrival Date': arrival,
 'Departure Date': departure,
 'Home Type': row['Home Type'],
 'Original Room Number': room,
 'Recommended Room Number': target_room,
 'Total Rate': row['Total Room Rate'],
 'Reason for Change': reason,
 'Cascade Step': cascade_step
 })

 current_revenue[(target_room, year)] + = row['Total Room Rate']
 changes_made = True

 occupancy = calculate_daily_occupancy(df, assigned_rooms)
 pass_number + = 1

 print("\n = = = Phase 1 complete = = = \n")
 def recompute_dynamic_target_revenue(df, assigned_rooms, room_info):
 # First → compute Total Current Revenue per home type
 current_home_type_revenue = {}
 for _, row in df.iterrows():
 # Skip if room type occupancy is already > 70%
 pass # placeholder block for for-loop
 total_nights = len(type_rooms)* 365
 occupancy = occupied_nights / total_nights
 if occupancy > 0.7:
 if available_nights = = 0:
 continue
 final_room = assigned_rooms.get(conf, row['Room Number'])
 home_type_final = room_info[final_room]['Room Type']
 key =(home_type_final, year)
 year = row['Year']
 current_home_type_revenue[key] = current_home_type_revenue.get(key, 0.0)+ row['Total Room Rate']

 # Second → compute Available Fraction per room(unchanged from original logic)
 target_revenue_list = []
 for room, info in room_info.items():
 room_type = info['Room Type']
 first_avail = info['First Available']

 for year in df['Year'].unique():
 days_in_year = 366 if year % 4 = = 0 else 365

 if first_avail.year > year:
 avail_fraction = 0.0
 elif first_avail.year = = year and first_avail > pd.Timestamp(f"{year}-01-01"):
 avail_fraction =(days_in_year -(first_avail - pd.Timestamp(f"{year}-01-01")).days)/ days_in_year
 else:
 avail_fraction = 1.0

 # Now compute Total Available Fraction for current rooms in this home type and year
 available_rooms = [
 r for r, i in room_info.items()
 if i['Room Type'] = = room_type
 and(i['First Available'].year < year or(i['First Available'].year = = year and i['First Available'] < = pd.Timestamp(f"{year}-01-01")))
 ]

 total_available_fraction = 0.0
 for r in available_rooms:
 first_avail_r = room_info[r]['First Available']
 if first_avail_r.year > year:
 f = 0.0
 elif first_avail_r.year = = year and first_avail_r > pd.Timestamp(f"{year}-01-01"):
 f =(days_in_year -(first_avail_r - pd.Timestamp(f"{year}-01-01")).days)/ days_in_year
 else:
 f = 1.0
 total_available_fraction + = f

 # Now compute Target Revenue for this room
 key =(room_type, year)
 total_current_revenue = current_home_type_revenue.get(key, 0.0)

 if total_available_fraction = = 0:
 target_rev = 0.0
 else:
 target_rev = avail_fraction * total_current_revenue / total_available_fraction

 target_revenue_list.append({
 'Room Number': room,
 'Year': year,
 'Target Revenue': target_rev
 })

 # Return as DataFrame → ready to use in Spread Revenue loop
 return pd.DataFrame(target_revenue_list)

 # = = = = = Final Phase 2 Loop: Spread Revenue with all improvements = = = = =

 print("\n = = = Phase 2 → Spread Revenue(< 70%)= = = ")
 df_sorted = df.copy()
 df_sorted['Nights'] =(pd.to_datetime(df_sorted['Departure Date'])- pd.to_datetime(df_sorted['Arrival Date'])).dt.days.replace(0, 1)
 df_sorted['SortKey'] = df_sorted['Total Room Rate'] / df_sorted['Nights']
 df_sorted = df_sorted.sort_values(by = 'SortKey', ascending = False)

 pass_number = 0
 changes_made = True

 while changes_made:
 print(f"\n--- Phase 2 → Pass {pass_number} ---")
 changes_made = False

 for _, row in df_sorted.iterrows():
 conf = row['Confirmation Number']
 arrival = row['Arrival Date']
 departure = row['Departure Date']
 room = row['Room Number']
 home_type = row['Room Type']
 do_not_move = str(row.get('Do Not Move', '')).upper()= = 'YES'
 is_comp = row['Rate Plan'] = = 'COMP'
 year = row['Year']

 if arrival < TODAY:
 continue

 current_room = assigned_rooms.get(conf, room)

 available_rooms = [r for r in room_info if room_info[r]['Room Type'] = = home_type]

 # Improved → Mean occupancy % instead of strict all()check
 occupancy_pct_list = [
(occupancy.get((home_type, d), {'Owner': 0, 'Comp': 0, 'Transient': 0})['Transient'] /
 max(1, len(available_rooms)-
 occupancy.get((home_type, d), {'Owner': 0, 'Comp': 0, 'Transient': 0})['Owner'] -
 occupancy.get((home_type, d), {'Owner': 0, 'Comp': 0, 'Transient': 0})['Comp']))
 for d in pd.date_range(arrival, departure - timedelta(days = 1)).date
 ]

 occupancy_below_threshold =(np.mean(occupancy_pct_list)< OCCUPANCY_THRESHOLD)

 reason = None
 target_room = None

 if(pd.isna(room)or room = = 'nan')and assigned_rooms.get(conf)is None:
 reason = 'Initial Room Assignment'
 elif do_not_move:
 assigned_rooms[conf] = current_room
 continue
 elif occupancy_below_threshold:
 reason = 'Spread Revenue'
 else:
 continue

 skip_rooms = set(r for r, info in room_info.items()if info['First Available'] > arrival)
 target_room = find_available_room(home_type, arrival, departure, is_comp, skip_rooms, reason, year)

 cascade_step = pass_number

 if target_room and target_room ! = current_room:
 for single_date in pd.date_range(arrival, departure - timedelta(days = 1)):
 room_calendar[target_room].add(single_date.date())
 assigned_rooms[conf] = target_room

 print(f"Moving CONF {conf}({reason})to room {target_room}(Cascade Step {cascade_step})")

 change_recommendations.append({
 'Confirmation Number': conf,
 'Arrival Date': arrival,
 'Departure Date': departure,
 'Home Type': row['Home Type'],
 'Original Room Number': room,
 'Recommended Room Number': target_room,
 'Total Rate': row['Total Room Rate'],
 'Reason for Change': reason,
 'Cascade Step': cascade_step
 })

 current_revenue[(target_room, year)] + = row['Total Room Rate']
 changes_made = True

 occupancy = calculate_daily_occupancy(df, assigned_rooms)

 # 🚀 Dynamic Target Revenue recompute after each pass
 target_revenue_df = recompute_dynamic_target_revenue(df, assigned_rooms, room_info)

 pass_number + = 1

 # Optional → shuffle stays each pass → avoids early stay bias
 df_sorted = df_sorted.sample(frac = 1, random_state = pass_number).reset_index(drop = True)

 print("\n = = = Phase 2 complete = = = \n")

 target_revenue_df = recompute_dynamic_target_revenue(df, assigned_rooms, room_info)

 print("\n = = = Phase 3 → Rebalance Overfilled Rooms to Underfilled Ones = = = ")

 # Recompute Target Revenue and Occupancy again(safe refresh)
 target_revenue_df = recompute_dynamic_target_revenue(df, assigned_rooms, room_info)
 occupancy = calculate_daily_occupancy(df, assigned_rooms)

 # Build Target and Current Revenue lookup
 target_dict = {(row['Room Number'], row['Year']): row['Target Revenue'] for _, row in target_revenue_df.iterrows()}

 # Classify overfilled and underfilled rooms
 overfilled = {}
 underfilled = {}
 for(room, year), target in target_dict.items():
 curr = sum(
 row['Total Room Rate'] for _, row in df.iterrows()
 if assigned_rooms.get(row['Confirmation Number'], row['Room Number'])= = room and row['Year'] = = year
)
 if target = = 0:
 continue
 pct = curr / target
 if pct > 1.10:
 overfilled[(room, year)] = pct
 elif pct < 0.90:
 underfilled[(room, year)] = pct

 rebalance_moves = 0

 # For each overfilled room, try to move one stay to underfilled rooms
 for(over_room, year), over_pct in sorted(overfilled.items(), key = lambda x: -x[1]):
 over_home_type = room_info[over_room]['Room Type']

 for _, row in df.iterrows():
 # Skip if room type occupancy is already > 70%
 occupied_nights = sum(len(room_calendar.get(r, set()))for r in type_rooms)
 total_nights = len(type_rooms)* 365
 occupancy = occupied_nights / total_nights
 if occupancy > 0.7:
 continue # completes `if occupancy > 0.7: `
 if available_nights = = 0:
 continue
 pass # placeholder to satisfy if block
 if assigned_rooms.get(conf, row['Room Number'])! = over_room:
 continue
 if row['Year'] ! = year:
 continue
 if str(row.get('Do Not Move', '')).upper()= = 'YES':
 continue
 if row['Rate Plan'] in owner_rate_plans or row['Rate Plan'] = = comp_rate_plan:
 continue

 arrival = row['Arrival Date']
 departure = row['Departure Date']
 is_comp = row['Rate Plan'] = = 'COMP'
 moved = False

 for(under_room, u_pct)in sorted(
 [(r, p)for(r, y), p in underfilled.items()
 if y = = year and room_info[r]['Room Type'] = = over_home_type],
 key = lambda x: x[1]
):
 if under_room = = over_room or room_info[under_room]['First Available'] > arrival:
 continue

 # Check for date conflicts
 room_dates = room_calendar[under_room]
 conflict = any(date in room_dates for date in get_stay_dates(arrival, departure))
 if not conflict:
 # ✅ Perform rebalancing move
 assigned_rooms[conf] = under_room
 for d in pd.date_range(arrival, departure - timedelta(days = 1)):
 room_calendar[under_room].add(d.date())
 rebalance_moves + = 1
 moved = True

 change_recommendations.append({
 'Confirmation Number': conf,
 'Arrival Date': arrival,
 'Departure Date': departure,
 'Home Type': row['Home Type'],
 'Original Room Number': over_room,
 'Recommended Room Number': under_room,
 'Total Rate': row['Total Room Rate'],
 'Reason for Change': 'Rebalance',
 'Cascade Step': 'Phase 3'
 })

 print(f"→ Rebalanced CONF {conf} from {over_room}({over_pct: .1%})→ {under_room}({u_pct: .1%})")
 # continue trying other stays in the same room to reduce overfill

 if not moved:
 print(f"✘ Could not rebalance CONF {conf} from {over_room} — no eligible underfilled room found.")

 print(f"\n = = = Phase 3 complete — {rebalance_moves} rebalancing moves made = = = \n")

 target_revenue_df = recompute_dynamic_target_revenue(df, assigned_rooms, room_info)

 # Fix: Convert Cascade Step to string to avoid type sorting error
 df_changes = pd.DataFrame(change_recommendations)
 df_changes['Cascade Step'] = df_changes['Cascade Step'].astype(str)

 # Keep only the last recommendation per confirmation number
 df_changes_final = df_changes.sort_values(by = 'Cascade Step').groupby('Confirmation Number', as_index = False).last()

 # Recompute Current Revenue → Final Revenue per room → based on final room assignments
 print("\n = = = Recomputing Final Revenue per Room = = = ")

 current_revenue = {(row['Room Number'], row['Year']): 0.0 for _, row in target_revenue_df.iterrows()}

 for _, row in df.iterrows():
 # Skip if room type occupancy is already > 70%
 occupied_nights = sum(len(room_calendar.get(r, set()))for r in type_rooms)
 total_nights = len(type_rooms)* 365
 occupancy = occupied_nights / total_nights
 if occupancy > 0.7:
 continue # completes `if occupancy > 0.7: `
 if available_nights = = 0:
 continue
 continue
 arrival = row['Arrival Date']
 departure = row['Departure Date']
 final_room = assigned_rooms.get(conf, row['Room Number'])
 year = row['Year']

 # Add full stay revenue to final_room's Current Revenue
 current_revenue[(final_room, year)] + = row['Total Room Rate']

 # Compute room revenue & occupancy
 room_stats = []
 for(room, year), target in target_dict.items():
 actual = current_revenue.get((room, year), 0.0)
 occupancy_nights = sum(
(row['Departure Date'] - row['Arrival Date']).days
 for _, row in df.iterrows()
 if assigned_rooms.get(row['Confirmation Number'], row['Room Number'])= = room and row['Year'] = = year
 and row['Rate Plan'] not in owner_rate_plans | {comp_rate_plan}
)

 available_nights = 365
 first_avail = room_info[room]['First Available']
 if first_avail.year = = year:
 available_nights =(pd.Timestamp(f"{year + 1}-01-01")- max(first_avail, pd.Timestamp(f"{year}-01-01"))).days
 elif first_avail.year > year:
 available_nights = 0

 original_occupancy_pct =(occupancy_nights / available_nights)if available_nights > 0 else 0
 actual_pct_of_target = actual / target if target > 0 else 0

 print(f"Room {room} | Year {year} | Target: ${target: , .2f} | Actual: ${actual: , .2f} | {actual_pct_of_target: .1%} of Target")

 room_stats.append({
 'Year': year,
 'Room Number': room,
 'Target Revenue': target,
 'Final Room Revenue': actual,
 'Final % of Target': actual_pct_of_target,
 'Final Occupancy %': original_occupancy_pct
 })

 room_stats_df = pd.DataFrame(room_stats)

 print("\n = = = Validating for overlapping stays(Auto-Fix Enabled)= = = ")
 conflict_count = 0
 fix_count = 0
 unfixable = 0
 room_night_tracker = {}
 unfixable_conflicts = []

 for _, row in df.iterrows():
 # Skip if room type occupancy is already > 70%
 occupied_nights = sum(len(room_calendar.get(r, set()))for r in type_rooms)
 total_nights = len(type_rooms)* 365
 occupancy = occupied_nights / total_nights
 if occupancy > 0.7:
 continue # completes `if occupancy > 0.7: `
 if available_nights = = 0:
 continue
 continue
 arrival = row['Arrival Date']
 departure = row['Departure Date']
 assigned_room = assigned_rooms.get(conf, row['Room Number'])
 home_type = row['Room Type']
 is_comp = row['Rate Plan'] = = comp_rate_plan
 year = row['Year']

 import pandas as pd
 OCCUPANCY_THRESHOLD = 0.7 # easy to adjust
 SPREAD_BUFFER = 0.90 # spread revenue buffer → rooms fill to 90% of target

 # Home type upgrade priority
 home_type_priority = ['B1QL', 'G4K', 'G4KP', 'G5K', 'G5KH', 'G5KO', 'G5KP', 'W5KO', 'W6KO', 'V6K', 'V6KR']

 # Room Type → Home Type mapping(normalized)

 UPGRADE_PATH = {
 'B1QL': ['G4K', 'G4KP'],
 'G4K': ['G4KP', 'G5K'],
 'G4KP': ['G5K'],
 'G5K': ['G5KH', 'G5KO', 'G5KP'],
 'G5KH': ['G5KO', 'G5KP', 'W5KO'],
 'G5KO': ['G5KP', 'W5KO'],
 'G5KP': ['W5KO'],
 'W5KO': ['W6KO'],
 'W6KO': ['V6K'],
 'V6K': ['V6KR']
 }

 ROOM_TO_HOME_TYPE = {
 'B1QL': 'Surf Bungalow',
 'G4K': 'Resort Cottage', 'G4KP': 'Resort Cottage', 'G5K': 'Resort Cottage',
 'G5KH': 'Resort Cottage', 'G5KO': 'Resort Cottage', 'G5KP': 'Resort Cottage',
 'V6K': 'Beach Villa', 'V6KR': 'Beach Villa',
 'W5KO': 'Waterside Retreat', 'W6KO': 'Waterside Retreat'
 }

 home_type_rank = {ht: i for i, ht in enumerate(home_type_priority)}
 get_stay_dates = lambda arrival, departure: [d.date()for d in pd.date_range(arrival, departure - timedelta(days = 1))]

 # Load data
 sheet1_df = pd.read_excel('Sheet1.xlsx')
 availability_df = pd.read_excel('Availability.xlsx')

 # Clean stays
 df = sheet1_df[~sheet1_df['Status'].isin(['Cancelled', 'No Show'])].copy()
 df['Arrival Date'] = pd.to_datetime(df['Arrival Date']).dt.normalize()
 df['Departure Date'] = pd.to_datetime(df['Departure Date']).dt.normalize()
 df['Room Type'] = df['Room Type'].astype(str)
 df['Room Number'] = df['Room'].astype(str)
 df['Home Type'] = df['Room Type'].map(ROOM_TO_HOME_TYPE)

 # = = = Explode Stays to Nightly Rate for Year-Based Targeting = = =
 expanded_rows = []
 for _, row in df.iterrows():
 stay_nights =(row['Departure Date'] - row['Arrival Date']).days
 if stay_nights = = 0:
 continue
 nightly_rate = row['Total Room Rate'] / stay_nights
 for single_date in pd.date_range(row['Arrival Date'], row['Departure Date'] - timedelta(days = 1)):
 expanded_rows.append({
 'Room Type': row['Room Type'],
 'Date': single_date.date(),
 'Total Rate': nightly_rate
 })

 expanded_df = pd.DataFrame(expanded_rows)
 expanded_df['Year'] = pd.to_datetime(expanded_df['Date']).dt.year
 expanded_df['Home Type'] = expanded_df['Room Type'].map(ROOM_TO_HOME_TYPE)

 # Process availability
 availability_df = availability_df[availability_df['Availability'] = = 'Available'].copy()
 availability_df['First Date of Availability'] = pd.to_datetime(availability_df['First Date of Availability']).dt.normalize()
 availability_df['Comp'] = availability_df['Comp'].fillna('').str.upper().eq('YES')

 # Build room availability dict
 room_info = {}
 for _, row in availability_df.iterrows():
 room_number = str(row['Homes'])
 room_type = row['Room Type']
 first_avail = row['First Date of Availability']
 is_comp = row['Comp']
 room_info[room_number] = {
 'Room Type': room_type,
 'Home Type': ROOM_TO_HOME_TYPE.get(room_type),
 'First Available': first_avail,
 'Comp': is_comp
 }

 # Build initial room calendar
 room_calendar = {room: set()for room in room_info}

 for _, row in df.iterrows():
 room = row['Room Number']
 if pd.notnull(room)and room ! = 'nan' and room in room_calendar:
 for single_date in pd.date_range(row['Arrival Date'], row['Departure Date'] - timedelta(days = 1)):
 room_calendar[room].add(single_date.date())

 # Owner and Comp Rate Plans
 owner_rate_plans = {'4RCOWN', 'BUNOWN', 'BVOWN', 'RCOWN', 'WROWN'}
 comp_rate_plan = 'COMP'

 # Build home type revenue per year
 df['Year'] = pd.to_datetime(df['Arrival Date']).dt.year
 home_type_year_revenue = expanded_df.groupby(['Room Type', 'Year'])['Total Rate'].sum().reset_index()

 # Build target revenue per room per year
 target_revenue = []
 for room, info in room_info.items():
 room_type = info['Room Type']
 first_avail = info['First Available']

 for year in df['Year'].unique():
 total_rev_row = home_type_year_revenue[(home_type_year_revenue['Room Type'] = = room_type)&(home_type_year_revenue['Year'] = = year)]
 if total_rev_row.empty:
 continue
 total_rev = total_rev_row['Total Rate'].values[0]

 days_in_year = 366 if year % 4 = = 0 else 365

 if first_avail.year > year:
 avail_fraction = 0.0
 elif first_avail.year = = year and first_avail > pd.Timestamp(f"{year}-01-01"):
 avail_fraction =(days_in_year -(first_avail - pd.Timestamp(f"{year}-01-01")).days)/ days_in_year
 else:
 avail_fraction = 1.0

 available_rooms_count = len([r for r, i in room_info.items()if i['Room Type'] = = room_type and(i['First Available'].year < year or(i['First Available'].year = = year and i['First Available'] < = pd.Timestamp(f"{year}-01-01")))])

 if available_rooms_count = = 0:
 target_rev = 0.0
 else:
 target_rev = avail_fraction * total_rev / available_rooms_count

 target_revenue.append({
 'Room Number': room,
 'Year': year,
 'Target Revenue': target_rev
 })

 target_revenue_df = pd.DataFrame(target_revenue)

 # Initialize current revenue tracker
 current_revenue = {(row['Room Number'], row['Year']): 0.0 for row in target_revenue}
 # Helper: calculate daily occupancy(unchanged)
 def calculate_daily_occupancy(df, assigned_rooms):
 occupancy = {}
 for _, row in df.iterrows():
 home_type = row['Room Type']
 room = assigned_rooms.get(row['Confirmation Number'], row['Room Number'])
 if pd.isna(room)or room = = 'nan':
 continue

 rate_plan = row['Rate Plan']
 for single_date in pd.date_range(row['Arrival Date'], row['Departure Date'] - timedelta(days = 1)):
 date = single_date.date()
 key =(home_type, date)
 if key not in occupancy:
 occupancy[key] = {'Owner': 0, 'Comp': 0, 'Transient': 0}

 if rate_plan in owner_rate_plans:
 occupancy[key]['Owner'] + = 1
 elif rate_plan = = comp_rate_plan:
 occupancy[key]['Comp'] + = 1
 else:
 occupancy[key]['Transient'] + = 1
 return occupancy

 def find_available_room(home_type, arrival, departure, is_comp, skip_rooms, reason, year):
 if reason = = 'Spread Revenue':
 # Only use same home type during Spread Revenue
 possible_rooms = [r for r, info in room_info.items()
 if info['Room Type'] = = home_type
 and(not is_comp or info['Comp'])
 and info['First Available'] < = arrival
 and r not in skip_rooms]

 room_target_pairs = []
 for room in possible_rooms:
 target_row = target_revenue_df[
(target_revenue_df['Room Number'] = = room)&(target_revenue_df['Year'] = = year)]
 if target_row.empty:
 continue
 target_rev = target_row['Target Revenue'].values[0]
 curr_rev = current_revenue.get((room, year), 0.0)

 if target_rev = = 0:
 continue

 pct_filled = curr_rev / target_rev
 if pct_filled > = SPREAD_BUFFER:
 continue # ✅ New: Don't allow rooms above the spread buffer

 room_target_pairs.append((room, pct_filled, curr_rev))

 # ✅ Sort by % of Target filled(asc), then by raw current revenue(asc)
 room_target_pairs.sort(key = lambda x: (x[1], x[2]))

 for room, pct_filled, _ in room_target_pairs:
 room_dates = room_calendar[room]
 conflict = any(date in room_dates for date in get_stay_dates(arrival, departure))
 if not conflict:
 return room

 else:
 # Consolidate phase → allow upgrades
 upgradeable_types = [home_type] + UPGRADE_PATH.get(home_type, [])
 possible_rooms = [r for r, info in room_info.items()
 if info['Room Type'] in upgradeable_types
 and(not is_comp or info['Comp'])
 and info['First Available'] < = arrival
 and r not in skip_rooms]

 possible_rooms = sorted(possible_rooms, key = lambda r: home_type_rank[room_info[r]['Room Type']])
 for room in possible_rooms:
 room_dates = room_calendar[room]
 conflict = any(date in room_dates for date in get_stay_dates(arrival, departure))
 if not conflict:
 return room

 return None
 # Phase 1 → Consolidate loop(≥ 70%)
 print("\n = = = Phase 1 → Consolidate(≥ 70%)= = = ")
 df_sorted = df.copy()
 df_sorted['Nights'] =(pd.to_datetime(df_sorted['Departure Date'])- pd.to_datetime(df_sorted['Arrival Date'])).dt.days.replace(0, 1)
 df_sorted['SortKey'] = df_sorted['Total Room Rate'] / df_sorted['Nights']
 df_sorted = df_sorted.sort_values(by = 'SortKey', ascending = False)

 assigned_rooms = {}
 change_recommendations = []
 cascade_steps = {}
 occupancy = calculate_daily_occupancy(df, assigned_rooms)

 pass_number = 0
 changes_made = True

 while changes_made:
 print(f"\n--- Phase 1 → Pass {pass_number} ---")
 changes_made = False

 for _, row in df_sorted.iterrows():
 conf = row['Confirmation Number']
 arrival = row['Arrival Date']
 departure = row['Departure Date']
 room = row['Room Number']
 home_type = row['Room Type']
 do_not_move = str(row.get('Do Not Move', '')).upper()= = 'YES'
 is_comp = row['Rate Plan'] = = 'COMP'
 year = row['Year']

 if arrival < TODAY:
 continue

 current_room = assigned_rooms.get(conf, room)

 available_rooms = [r for r in room_info if room_info[r]['Room Type'] = = home_type]

 over_occupancy = any(
(occupancy.get((home_type, d), {'Owner': 0, 'Comp': 0, 'Transient': 0})['Transient'] /
 max(1, len(available_rooms)-
 occupancy.get((home_type, d), {'Owner': 0, 'Comp': 0, 'Transient': 0})['Owner'] -
 occupancy.get((home_type, d), {'Owner': 0, 'Comp': 0, 'Transient': 0})['Comp']))> OCCUPANCY_THRESHOLD
 for d in pd.date_range(arrival, departure - timedelta(days = 1)).date)

 reason = None
 target_room = None

 if(pd.isna(room)or room = = 'nan')and assigned_rooms.get(conf)is None:
 reason = 'Initial Room Assignment'
 elif do_not_move:
 assigned_rooms[conf] = current_room
 continue
 elif over_occupancy:
 reason = 'Consolidate'
 else:
 continue # Skip → Phase 1 only moves for Consolidate

 skip_rooms = set(r for r, info in room_info.items()if info['First Available'] > arrival)
 target_room = find_available_room(home_type, arrival, departure, is_comp, skip_rooms, reason, year)

 cascade_step = pass_number

 if target_room and target_room ! = current_room:
 for single_date in pd.date_range(arrival, departure - timedelta(days = 1)):
 room_calendar[target_room].add(single_date.date())
 assigned_rooms[conf] = target_room

 print(f"Moving CONF {conf}({reason})to room {target_room}(Cascade Step {cascade_step})")

 change_recommendations.append({
 'Confirmation Number': conf,
 'Arrival Date': arrival,
 'Departure Date': departure,
 'Home Type': row['Home Type'],
 'Original Room Number': room,
 'Recommended Room Number': target_room,
 'Total Rate': row['Total Room Rate'],
 'Reason for Change': reason,
 'Cascade Step': cascade_step
 })

 current_revenue[(target_room, year)] + = row['Total Room Rate']
 changes_made = True

 occupancy = calculate_daily_occupancy(df, assigned_rooms)
 pass_number + = 1

 print("\n = = = Phase 1 complete = = = \n")
 def recompute_dynamic_target_revenue(df, assigned_rooms, room_info):
 # First → compute Total Current Revenue per home type
 current_home_type_revenue = {}
 for _, row in df.iterrows():
 # Skip if room type occupancy is already > 70%
 occupied_nights = sum(len(room_calendar.get(r, set()))for r in type_rooms)
 total_nights = len(type_rooms)* 365
 occupancy = occupied_nights / total_nights
 if occupancy > 0.7:
 continue # completes `if occupancy > 0.7: `
 if available_nights = = 0:
 continue
 continue
 final_room = assigned_rooms.get(conf, row['Room Number'])
 home_type_final = room_info[final_room]['Room Type']
 year = row['Year']

 key =(home_type_final, year)
 current_home_type_revenue[key] = current_home_type_revenue.get(key, 0.0)+ row['Total Room Rate']

 # Second → compute Available Fraction per room(unchanged from original logic)
 target_revenue_list = []
 for room, info in room_info.items():
 room_type = info['Room Type']
 first_avail = info['First Available']

 for year in df['Year'].unique():
 days_in_year = 366 if year % 4 = = 0 else 365

 if first_avail.year > year:
 avail_fraction = 0.0
 elif first_avail.year = = year and first_avail > pd.Timestamp(f"{year}-01-01"):
 avail_fraction =(days_in_year -(first_avail - pd.Timestamp(f"{year}-01-01")).days)/ days_in_year
 else:
 avail_fraction = 1.0

 # Now compute Total Available Fraction for current rooms in this home type and year
 available_rooms = [
 r for r, i in room_info.items()
 if i['Room Type'] = = room_type
 and(i['First Available'].year < year or(i['First Available'].year = = year and i['First Available'] < = pd.Timestamp(f"{year}-01-01")))
 ]

 total_available_fraction = 0.0
 for r in available_rooms:
 first_avail_r = room_info[r]['First Available']
 if first_avail_r.year > year:
 f = 0.0
 elif first_avail_r.year = = year and first_avail_r > pd.Timestamp(f"{year}-01-01"):
 f =(days_in_year -(first_avail_r - pd.Timestamp(f"{year}-01-01")).days)/ days_in_year
 else:
 f = 1.0
 total_available_fraction + = f

 # Now compute Target Revenue for this room
 key =(room_type, year)
 total_current_revenue = current_home_type_revenue.get(key, 0.0)

 if total_available_fraction = = 0:
 target_rev = 0.0
 else:
 target_rev = avail_fraction * total_current_revenue / total_available_fraction

 target_revenue_list.append({
 'Room Number': room,
 'Year': year,
 'Target Revenue': target_rev
 })

 # Return as DataFrame → ready to use in Spread Revenue loop
 return pd.DataFrame(target_revenue_list)

 # = = = = = Final Phase 2 Loop: Spread Revenue with all improvements = = = = =

 print("\n = = = Phase 2 → Spread Revenue(< 70%)= = = ")
 df_sorted = df.copy()
 df_sorted['Nights'] =(pd.to_datetime(df_sorted['Departure Date'])- pd.to_datetime(df_sorted['Arrival Date'])).dt.days.replace(0, 1)
 df_sorted['SortKey'] = df_sorted['Total Room Rate'] / df_sorted['Nights']
 df_sorted = df_sorted.sort_values(by = 'SortKey', ascending = False)

 pass_number = 0
 changes_made = True

 while changes_made:
 print(f"\n--- Phase 2 → Pass {pass_number} ---")
 changes_made = False

 for _, row in df_sorted.iterrows():
 conf = row['Confirmation Number']
 arrival = row['Arrival Date']
 departure = row['Departure Date']
 room = row['Room Number']
 home_type = row['Room Type']
 do_not_move = str(row.get('Do Not Move', '')).upper()= = 'YES'
 is_comp = row['Rate Plan'] = = 'COMP'
 year = row['Year']

 if arrival < TODAY:
 continue

 current_room = assigned_rooms.get(conf, room)

 available_rooms = [r for r in room_info if room_info[r]['Room Type'] = = home_type]

 # Improved → Mean occupancy % instead of strict all()check
 occupancy_pct_list = [
(occupancy.get((home_type, d), {'Owner': 0, 'Comp': 0, 'Transient': 0})['Transient'] /
 max(1, len(available_rooms)-
 occupancy.get((home_type, d), {'Owner': 0, 'Comp': 0, 'Transient': 0})['Owner'] -
 occupancy.get((home_type, d), {'Owner': 0, 'Comp': 0, 'Transient': 0})['Comp']))
 for d in pd.date_range(arrival, departure - timedelta(days = 1)).date
 ]

 occupancy_below_threshold =(np.mean(occupancy_pct_list)< OCCUPANCY_THRESHOLD)

 reason = None
 target_room = None

 if(pd.isna(room)or room = = 'nan')and assigned_rooms.get(conf)is None:
 reason = 'Initial Room Assignment'
 elif do_not_move:
 assigned_rooms[conf] = current_room
 continue
 elif occupancy_below_threshold:
 reason = 'Spread Revenue'
 else:
 continue

 skip_rooms = set(r for r, info in room_info.items()if info['First Available'] > arrival)
 target_room = find_available_room(home_type, arrival, departure, is_comp, skip_rooms, reason, year)

 cascade_step = pass_number

 if target_room and target_room ! = current_room:
 for single_date in pd.date_range(arrival, departure - timedelta(days = 1)):
 room_calendar[target_room].add(single_date.date())
 assigned_rooms[conf] = target_room

 print(f"Moving CONF {conf}({reason})to room {target_room}(Cascade Step {cascade_step})")

 change_recommendations.append({
 'Confirmation Number': conf,
 'Arrival Date': arrival,
 'Departure Date': departure,
 'Home Type': row['Home Type'],
 'Original Room Number': room,
 'Recommended Room Number': target_room,
 'Total Rate': row['Total Room Rate'],
 'Reason for Change': reason,
 'Cascade Step': cascade_step
 })

 current_revenue[(target_room, year)] + = row['Total Room Rate']
 changes_made = True

 occupancy = calculate_daily_occupancy(df, assigned_rooms)

 # 🚀 Dynamic Target Revenue recompute after each pass
 target_revenue_df = recompute_dynamic_target_revenue(df, assigned_rooms, room_info)

 pass_number + = 1

 # Optional → shuffle stays each pass → avoids early stay bias
 df_sorted = df_sorted.sample(frac = 1, random_state = pass_number).reset_index(drop = True)

 print("\n = = = Phase 2 complete = = = \n")

 target_revenue_df = recompute_dynamic_target_revenue(df, assigned_rooms, room_info)

 print("\n = = = Phase 3 → Rebalance Overfilled Rooms to Underfilled Ones = = = ")

 # Recompute Target Revenue and Occupancy again(safe refresh)
 target_revenue_df = recompute_dynamic_target_revenue(df, assigned_rooms, room_info)
 occupancy = calculate_daily_occupancy(df, assigned_rooms)

 # Build Target and Current Revenue lookup
 target_dict = {(row['Room Number'], row['Year']): row['Target Revenue'] for _, row in target_revenue_df.iterrows()}

 # Classify overfilled and underfilled rooms
 overfilled = {}
 underfilled = {}
 for(room, year), target in target_dict.items():
 curr = sum(
 row['Total Room Rate'] for _, row in df.iterrows()
 if assigned_rooms.get(row['Confirmation Number'], row['Room Number'])= = room and row['Year'] = = year
)
 if target = = 0:
 continue
 pct = curr / target
 if pct > 1.10:
 overfilled[(room, year)] = pct
 elif pct < 0.90:
 underfilled[(room, year)] = pct

 rebalance_moves = 0

 # For each overfilled room, try to move one stay to underfilled rooms
 for(over_room, year), over_pct in sorted(overfilled.items(), key = lambda x: -x[1]):
 over_home_type = room_info[over_room]['Room Type']

 for _, row in df.iterrows():
 # Skip if room type occupancy is already > 70%
 occupied_nights = sum(len(room_calendar.get(r, set()))for r in type_rooms)
 total_nights = len(type_rooms)* 365
 occupancy = occupied_nights / total_nights
 if occupancy > 0.7:
 continue # completes `if occupancy > 0.7: `
 if available_nights = = 0:
 continue
 continue
 if assigned_rooms.get(conf, row['Room Number'])! = over_room:
 continue # completes `if room not in current_home_type_rooms`
 if assigned_rooms.get((room, arrival))= = 'X':
 continue
 continue
 continue
 if str(row.get('Do Not Move', '')).upper()= = 'YES':
 continue # completes `if room not in room_type_mapping`
 if str(row.get('Assigned Room Type', ''))= = '':
 continue
 continue
 continue

 arrival = row['Arrival Date']
 departure = row['Departure Date']
 is_comp = row['Rate Plan'] = = 'COMP'
 moved = False

 for(under_room, u_pct)in sorted(reversed(upgrade_priority), key = lambda x: x[1]):
 continue

 # Check for date conflicts
 room_dates = room_calendar[under_room]
 conflict = any(date in room_dates for date in get_stay_dates(arrival, departure))
 if not conflict:
 # ✅ Perform rebalancing move
 assigned_rooms[conf] = under_room
 for d in pd.date_range(arrival, departure - timedelta(days = 1)):
 room_calendar[under_room].add(d.date())
 rebalance_moves + = 1
 moved = True

 change_recommendations.append({
 'Confirmation Number': conf,
 'Arrival Date': arrival,
 'Departure Date': departure,
 'Home Type': row['Home Type'],
 'Original Room Number': over_room,
 'Recommended Room Number': under_room,
 'Total Rate': row['Total Room Rate'],
 'Reason for Change': 'Rebalance',
 'Cascade Step': 'Phase 3'
 })

 print(f"→ Rebalanced CONF {conf} from {over_room}({over_pct: .1%})→ {under_room}({u_pct: .1%})")
 # continue trying other stays in the same room to reduce overfill

 if not moved:
 continue
 print(f"\n = = = Phase 3 complete — {rebalance_moves} rebalancing moves made = = = \n")

 target_revenue_df = recompute_dynamic_target_revenue(df, assigned_rooms, room_info)

 # Fix: Convert Cascade Step to string to avoid type sorting error
 df_changes = pd.DataFrame(change_recommendations)
 df_changes['Cascade Step'] = df_changes['Cascade Step'].astype(str)

 # Keep only the last recommendation per confirmation number
 df_changes_final = df_changes.sort_values(by = 'Cascade Step').groupby('Confirmation Number', as_index = False).last()

 # Recompute Current Revenue → Final Revenue per room → based on final room assignments
 print("\n = = = Recomputing Final Revenue per Room = = = ")

 current_revenue = {(row['Room Number'], row['Year']): 0.0 for _, row in target_revenue_df.iterrows()}

 for _, row in df.iterrows():
 # Skip if room type occupancy is already > 70%
 occupied_nights = sum(len(room_calendar.get(r, set()))for r in type_rooms)
 total_nights = len(type_rooms)* 365
 occupancy = occupied_nights / total_nights
 if occupancy > 0.7:
 continue # completes `if occupancy > 0.7: `
 if assigned_rooms.get((room, single_date))= = 'X':
 continue
 arrival = row['Arrival Date']
 departure = row['Departure Date']
 final_room = assigned_rooms.get(conf, row['Room Number'])
 year = row['Year']

 # Add full stay revenue to final_room's Current Revenue
 current_revenue[(final_room, year)] + = row['Total Room Rate']

 # Compute room revenue & occupancy
 room_stats = []
 for(room, year), target in target_dict.items():
 actual = current_revenue.get((room, year), 0.0)
 occupancy_nights = sum(
(row['Departure Date'] - row['Arrival Date']).days
 for _, row in df.iterrows()
 if assigned_rooms.get(row['Confirmation Number'], row['Room Number'])= = room and row['Year'] = = year
 and row['Rate Plan'] not in owner_rate_plans | {comp_rate_plan}
)

 available_nights = 365
 first_avail = room_info[room]['First Available']
 if first_avail.year = = year:
 available_nights =(pd.Timestamp(f"{year + 1}-01-01")- max(first_avail, pd.Timestamp(f"{year}-01-01"))).days
 elif first_avail.year > year:
 available_nights = 0

 original_occupancy_pct =(occupancy_nights / available_nights)if available_nights > 0 else 0
 actual_pct_of_target = actual / target if target > 0 else 0

 print(f"Room {room} | Year {year} | Target: ${target: , .2f} | Actual: ${actual: , .2f} | {actual_pct_of_target: .1%} of Target")

 room_stats.append({
 'Year': year,
 'Room Number': room,
 'Target Revenue': target,
 'Final Room Revenue': actual,
 'Final % of Target': actual_pct_of_target,
 'Final Occupancy %': original_occupancy_pct
 })

 room_stats_df = pd.DataFrame(room_stats)

 print("\n = = = Validating for overlapping stays(Auto-Fix Enabled)= = = ")
 conflict_count = 0
 fix_count = 0
 unfixable = 0
 room_night_tracker = {}
 unfixable_conflicts = []

 for _, row in df.iterrows():
 # Skip if room type occupancy is already > 70%
 occupied_nights = sum(len(room_calendar.get(r, set()))for r in type_rooms)
 total_nights = len(type_rooms)* 365
 occupancy = occupied_nights / total_nights
 if occupancy > 0.7:
 if not pd.isna(conf):
 pass
 continue # Placeholder to maintain structure
 if assigned_room not in room_night_tracker:
 room_night_tracker[assigned_room] = {}
 for d in pd.date_range(arrival, departure - timedelta(days = 1)).date:
 if d in room_night_tracker[assigned_room]:
 has_conflict = True
 break
 break

 if not has_conflict:
 # No conflict, register dates
 for d in pd.date_range(arrival, departure - timedelta(days = 1)).date:
 room_night_tracker[assigned_room][d] = conf
 continue

 # 🧠 Conflict detected — try to reassign
 # 🧠 Conflict detected — try to reassign
 conflict_count + = 1
 skip_rooms = set(r for r, info in room_info.items()if info['First Available'] > arrival)
 new_room = find_available_room(home_type, arrival, departure, is_comp, skip_rooms, reason = 'Conflict Fix', year = year)
 if new_room:
 fix_count + = 1
 assigned_rooms[conf] = new_room
 # Update new room's calendar
 if new_room not in room_night_tracker:
 room_night_tracker[new_room] = {}
 for d in pd.date_range(arrival, departure - timedelta(days = 1)).date:
 room_night_tracker[new_room][d] = conf

 change_recommendations.append({
 'Confirmation Number': conf,
 'Arrival Date': arrival,
 'Departure Date': departure,
 'Home Type': row['Home Type'],
 'Original Room Number': assigned_room,
 'Recommended Room Number': new_room,
 'Total Rate': row['Total Room Rate'],
 'Reason for Change': 'Conflict Auto-Fix',
 'Cascade Step': 'Validation'
 })
 if new_room:
 fix_count + = 1
 assigned_rooms[conf] = new_room
 print(f"✅ Conflict fixed: Moved CONF {conf} → Room {new_room}")
 else:
 unfixable + = 1
 print(f"❌ Could not resolve conflict for CONF {conf}(Room {assigned_room})")
 assignment_detail = {
 'Arrival Date': arrival,
 'Departure Date': departure,
 'Original Room Number': assigned_room,
 'Home Type': row['Home Type'],
 }
 conflict_report.append({
 'Total Rate': row['Total Room Rate'],
 'Conflict Date Range': f"{arrival} to {departure - timedelta(days = 1)}"
 })
 print(f"\n✅ {fix_count} conflicts fixed automatically.")
 if unfixable:
 print(f"❌ {unfixable} could not be auto-fixed — manual review may be needed.")

 # Write all outputs to Excel
 with pd.ExcelWriter(os.path.join(OUTPUT_FOLDER, OUTPUT_FILENAME), engine = 'xlsxwriter')as writer:
 # = = = Formatting applied to each sheet = = =
 for sheet in writer.sheets:
 worksheet = writer.sheets[sheet]
 df_sheet = eval(f"df_changes_final if sheet = = 'Recommended Changes' else room_stats_df if sheet = = 'Room Revenue & Occupancy' else pd.DataFrame(unfixable_conflicts)")
 for i, col in enumerate(df_sheet.columns, 1):
 max_length = max([len(str(cell))for cell in [col] + df_sheet[col].astype(str).tolist()])
 worksheet.set_column(i-1, i-1, max_length + 2)

 # Apply specific number formatting
 if sheet = = 'Recommended Changes':
 total_rate_idx = df_sheet.columns.get_loc('Total Rate')
 worksheet.set_column(total_rate_idx, total_rate_idx, None, writer.book.add_format({'num_format': '#, ##0.00'}))

 elif sheet = = 'Room Revenue & Occupancy':
 format_currency = writer.book.add_format({'num_format': '#, ##0.00'})
 format_percent = writer.book.add_format({'num_format': '0.00%'})
 for col_name in ['Target Revenue', 'Final Room Revenue']:
 if col_name in df_sheet.columns:
 idx = df_sheet.columns.get_loc(col_name)
 worksheet.set_column(idx, idx, None, format_currency)
 for col_name in ['Final % of Target', 'Final Occupancy %']:
 if col_name in df_sheet.columns:
 idx = df_sheet.columns.get_loc(col_name)
 worksheet.set_column(idx, idx, None, format_percent)

 elif sheet = = 'Unfixable Conflicts':
 if 'Total Rate' in df_sheet.columns:
 idx = df_sheet.columns.get_loc('Total Rate')
 worksheet.set_column(idx, idx, None, writer.book.add_format({'num_format': '#, ##0.00'}))

 df_changes_final.to_excel(writer, sheet_name = 'Recommended Changes', index = False)
 room_stats_df.to_excel(writer, sheet_name = 'Room Revenue & Occupancy', index = False)

 # Autofit columns for 'Recommended Changes'
 worksheet1 = writer.sheets['Recommended Changes']
 for i, col in enumerate(df_changes_final.columns):
 max_len = max(df_changes_final[col].astype(str).map(len).max(), len(str(col)))
 worksheet1.set_column(i, i, max_len + 2)

 # Autofit columns for 'Room Revenue & Occupancy'
 worksheet2 = writer.sheets['Room Revenue & Occupancy']
 for i, col in enumerate(room_stats_df.columns):
 max_len = max(room_stats_df[col].astype(str).map(len).max(), len(str(col)))
 worksheet2.set_column(i, i, max_len + 2)

 if unfixable_conflicts:
 pd.DataFrame(unfixable_conflicts).to_excel(writer, sheet_name = 'Unfixable Conflicts', index = False)

 print(f"\n✅ Exported results to: {os.path.join(OUTPUT_FOLDER, OUTPUT_FILENAME)}")

 # = = = Add Availability Charts to Excel Sheet = = =
 from openpyxl import load_workbook
 from openpyxl.drawing.image import Image as XLImage
 import glob
 import time

 try:
 full_path = os.path.join(OUTPUT_FOLDER, OUTPUT_FILENAME)

 # Wait briefly to ensure Excel file is fully written
 if not os.path.exists(full_path):
 raise FileNotFoundError(f"Excel output not found at {full_path}")

 time.sleep(2)# Give system time to release file handles
 wb = load_workbook(full_path)

 chart_sheet = wb.create_sheet("Availability Charts")

 row = 1
 chart_paths = sorted(glob.glob(os.path.join(OUTPUT_FOLDER, "Availability_Comparison_*.png")))

 if not chart_paths:
 print("⚠️ No availability charts found to insert into Excel.")

 for chart_path in chart_paths:
 if os.path.exists(chart_path):
 img = XLImage(chart_path)
 img.anchor = f"A{row}"
 chart_sheet.add_image(img)
 row + = 35
 else:
 print(f"⚠️ Missing chart image: {chart_path}")

 final_output_path = full_path.replace(".xlsx", " - Charts.xlsx")
 wb.save(final_output_path)
 print(f"✅ Charts successfully embedded in Excel: {final_output_path}")

 except Exception as e:
 print(f"⚠️ Error embedding charts in Excel: {e}")

 from openpyxl import load_workbook
 from openpyxl.drawing.image import Image as XLImage

 try:
 full_path = os.path.join(OUTPUT_FOLDER, OUTPUT_FILENAME)
 wb = load_workbook(full_path)

 chart_sheet = wb.create_sheet("Availability Charts")

 import glob
 row = 1
 chart_paths = glob.glob(os.path.join(OUTPUT_FOLDER, "Availability_Comparison_*.png"))

 for chart_path in sorted(chart_paths):
 img = XLImage(chart_path)
 img.anchor = f"A{row}"
 chart_sheet.add_image(img)
 row + = 35 # Space between images

 final_output_path = full_path.replace(".xlsx", " - Charts.xlsx")
 wb.save(final_output_path)
 print(f"✅ Charts added to Excel: {final_output_path}")
 except Exception as e:
 print(f"⚠️ Error embedding charts in Excel: {e}")
 has_conflict = True
 if not has_conflict:
 # No conflict, register dates
 for d in pd.date_range(arrival, departure - timedelta(days = 1)).date:
 room_night_tracker[assigned_room][d] = conf
 # continue # removed or to be placed inside loop
 # 🧠 Conflict detected — try to reassign
 conflict_count + = 1
 try:
 conflict_count + = 1
 skip_rooms = set(r for r, info in room_info.items()if info['First Available'] > arrival)
 new_room = find_available_room(home_type, arrival, departure, is_comp, skip_rooms, reason = 'Conflict Fix', year = year)
 if new_room:
 print(f"✅ Conflict fixed: Moved CONF {conf} → Room {new_room}")
 else:
 unfixable + = 1
 except Exception as e:
 print(f"⚠️ Conflict reassignment error: {e}")
 assigned_rooms[conf] = new_room

 # Update new room's calendar
 if new_room not in room_night_tracker:
 room_night_tracker[new_room] = {}
 for d in pd.date_range(arrival, departure - timedelta(days = 1)).date:
 room_night_tracker[new_room][d] = conf

 change_recommendations.append({
 'Confirmation Number': conf,
 'Arrival Date': arrival,
 'Departure Date': departure,
 'Home Type': row['Home Type'],
 'Original Room Number': assigned_room,
 'Recommended Room Number': new_room,
 'Total Rate': row['Total Room Rate'],
 'Reason for Change': 'Conflict Auto-Fix',
 'Cascade Step': 'Validation'
 })

 print(f"✅ Conflict fixed: Moved CONF {conf} → Room {new_room}")
 else:
 unfixable + = 1
 print(f"❌ Could not resolve conflict for CONF {conf}(Room {assigned_room})")
 unfixable_conflicts.append({
 'Confirmation Number': conf,
 'Arrival Date': arrival,
 'Departure Date': departure,
 'Original Room Number': assigned_room,
 'Home Type': row['Home Type'],
 'Rate Plan': row['Rate Plan'],
 'Total Rate': row['Total Room Rate'],
 'Conflict Date Range': f"{arrival} to {departure - timedelta(days = 1)}"
 })

 print(f"\n✅ {fix_count} conflicts fixed automatically.")
 if unfixable:
 print(f"❌ {unfixable} could not be auto-fixed — manual review may be needed.")

 # Write all outputs to Excel
 with pd.ExcelWriter(os.path.join(OUTPUT_FOLDER, OUTPUT_FILENAME), engine = 'xlsxwriter')as writer:
 # = = = Formatting applied to each sheet = = =
 for sheet in writer.sheets:
 worksheet = writer.sheets[sheet]
 df_sheet = eval(f"df_changes_final if sheet = = 'Recommended Changes' else room_stats_df if sheet = = 'Room Revenue & Occupancy' else pd.DataFrame(unfixable_conflicts)")
 for i, col in enumerate(df_sheet.columns, 1):
 max_length = max([len(str(cell))for cell in [col] + df_sheet[col].astype(str).tolist()])
 worksheet.set_column(i-1, i-1, max_length + 2)

 # Apply specific number formatting
 if sheet = = 'Recommended Changes':
 total_rate_idx = df_sheet.columns.get_loc('Total Rate')
 worksheet.set_column(total_rate_idx, total_rate_idx, None, writer.book.add_format({'num_format': '#, ##0.00'}))

 elif sheet = = 'Room Revenue & Occupancy':
 format_currency = writer.book.add_format({'num_format': '#, ##0.00'})
 format_percent = writer.book.add_format({'num_format': '0.00%'})
 for col_name in ['Target Revenue', 'Final Room Revenue']:
 if col_name in df_sheet.columns:
 idx = df_sheet.columns.get_loc(col_name)
 worksheet.set_column(idx, idx, None, format_currency)
 for col_name in ['Final % of Target', 'Final Occupancy %']:
 if col_name in df_sheet.columns:
 idx = df_sheet.columns.get_loc(col_name)
 worksheet.set_column(idx, idx, None, format_percent)

 elif sheet = = 'Unfixable Conflicts':
 if 'Total Rate' in df_sheet.columns:
 idx = df_sheet.columns.get_loc('Total Rate')
 worksheet.set_column(idx, idx, None, writer.book.add_format({'num_format': '#, ##0.00'}))

 df_changes_final.to_excel(writer, sheet_name = 'Recommended Changes', index = False)
 room_stats_df.to_excel(writer, sheet_name = 'Room Revenue & Occupancy', index = False)

 # Autofit columns for 'Recommended Changes'
 worksheet1 = writer.sheets['Recommended Changes']
 for i, col in enumerate(df_changes_final.columns):
 max_len = max(df_changes_final[col].astype(str).map(len).max(), len(str(col)))
 worksheet1.set_column(i, i, max_len + 2)

 # Autofit columns for 'Room Revenue & Occupancy'
 worksheet2 = writer.sheets['Room Revenue & Occupancy']
 for i, col in enumerate(room_stats_df.columns):
 max_len = max(room_stats_df[col].astype(str).map(len).max(), len(str(col)))
 worksheet2.set_column(i, i, max_len + 2)

 if unfixable_conflicts:
 pd.DataFrame(unfixable_conflicts).to_excel(writer, sheet_name = 'Unfixable Conflicts', index = False)

 print(f"\n✅ Exported results to: {os.path.join(OUTPUT_FOLDER, OUTPUT_FILENAME)}")

 # = = = Add Availability Charts to Excel Sheet = = =
 from openpyxl import load_workbook
 from openpyxl.drawing.image import Image as XLImage
 import glob
 import time

 try:
 full_path = os.path.join(OUTPUT_FOLDER, OUTPUT_FILENAME)

 # Wait briefly to ensure Excel file is fully written
 if not os.path.exists(full_path):
 raise FileNotFoundError(f"Excel output not found at {full_path}")

 time.sleep(2)# Give system time to release file handles
 wb = load_workbook(full_path)

 chart_sheet = wb.create_sheet("Availability Charts")

 row = 1
 chart_paths = sorted(glob.glob(os.path.join(OUTPUT_FOLDER, "Availability_Comparison_*.png")))

 if not chart_paths:
 print("⚠️ No availability charts found to insert into Excel.")

 for chart_path in chart_paths:
 if os.path.exists(chart_path):
 img = XLImage(chart_path)
 img.anchor = f"A{row}"
 chart_sheet.add_image(img)
 row + = 35
 else:
 print(f"⚠️ Missing chart image: {chart_path}")

 final_output_path = full_path.replace(".xlsx", " - Charts.xlsx")
 wb.save(final_output_path)
 print(f"✅ Charts successfully embedded in Excel: {final_output_path}")

 except Exception as e:
 print(f"⚠️ Error embedding charts in Excel: {e}")

 from openpyxl import load_workbook
 from openpyxl.drawing.image import Image as XLImage

 try:
 full_path = os.path.join(OUTPUT_FOLDER, OUTPUT_FILENAME)
 wb = load_workbook(full_path)

 chart_sheet = wb.create_sheet("Availability Charts")

 import glob
 row = 1
 chart_paths = glob.glob(os.path.join(OUTPUT_FOLDER, "Availability_Comparison_*.png"))

 for chart_path in sorted(chart_paths):
 img = XLImage(chart_path)
 img.anchor = f"A{row}"
 chart_sheet.add_image(img)
 row + = 35 # Space between images

 final_output_path = full_path.replace(".xlsx", " - Charts.xlsx")
 wb.save(final_output_path)
 print(f"✅ Charts added to Excel: {final_output_path}")
 except Exception as e:
 print(f"⚠️ Error embedding charts in Excel: {e}")
 has_conflict = True
 if not has_conflict:
 # No conflict, register dates
 for d in pd.date_range(arrival, departure - timedelta(days = 1)).date:
 room_night_tracker[assigned_room][d] = conf
 continue # Validated loop-level continue
 # 🧠 Conflict detected — try to reassign
 conflict_count + = 1
 skip_rooms = set(r for r, info in room_info.items()if info['First Available'] > arrival)
 new_room = find_available_room(home_type, arrival, departure, is_comp, skip_rooms, reason = 'Conflict Fix', year = year)

 if new_room:
 fix_count + = 1
 assigned_rooms[conf] = new_room

 # Update new room's calendar
 if new_room not in room_night_tracker:
 room_night_tracker[new_room] = {}
 for d in pd.date_range(arrival, departure - timedelta(days = 1)).date:
 room_night_tracker[new_room][d] = conf

 change_recommendations.append({
 'Confirmation Number': conf,
 'Arrival Date': arrival,
 'Departure Date': departure,
 'Home Type': row['Home Type'],
 'Original Room Number': assigned_room,
 'Recommended Room Number': new_room,
 'Total Rate': row['Total Room Rate'],
 'Reason for Change': 'Conflict Auto-Fix',
 'Cascade Step': 'Validation'
 })

 print(f"✅ Conflict fixed: Moved CONF {conf} → Room {new_room}")
 else:
 unfixable + = 1
 print(f"❌ Could not resolve conflict for CONF {conf}(Room {assigned_room})")
 unfixable_conflicts.append({
 'Confirmation Number': conf,
 'Arrival Date': arrival,
 'Departure Date': departure,
 'Original Room Number': assigned_room,
 'Home Type': row['Home Type'],
 'Rate Plan': row['Rate Plan'],
 'Total Rate': row['Total Room Rate'],
 'Conflict Date Range': f"{arrival} to {departure - timedelta(days = 1)}"
 })

 print(f"\n✅ {fix_count} conflicts fixed automatically.")
 if unfixable:
 print(f"❌ {unfixable} could not be auto-fixed — manual review may be needed.")

 # Write all outputs to Excel
 with pd.ExcelWriter(os.path.join(OUTPUT_FOLDER, OUTPUT_FILENAME), engine = 'xlsxwriter')as writer:
 # = = = Formatting applied to each sheet = = =
 for sheet in writer.sheets:
 worksheet = writer.sheets[sheet]
 df_sheet = eval(f"df_changes_final if sheet = = 'Recommended Changes' else room_stats_df if sheet = = 'Room Revenue & Occupancy' else pd.DataFrame(unfixable_conflicts)")
 for i, col in enumerate(df_sheet.columns, 1):
 max_length = max([len(str(cell))for cell in [col] + df_sheet[col].astype(str).tolist()])
 worksheet.set_column(i-1, i-1, max_length + 2)

 # Apply specific number formatting
 if sheet = = 'Recommended Changes':
 total_rate_idx = df_sheet.columns.get_loc('Total Rate')
 worksheet.set_column(total_rate_idx, total_rate_idx, None, writer.book.add_format({'num_format': '#, ##0.00'}))

 elif sheet = = 'Room Revenue & Occupancy':
 format_currency = writer.book.add_format({'num_format': '#, ##0.00'})
 format_percent = writer.book.add_format({'num_format': '0.00%'})
 for col_name in ['Target Revenue', 'Final Room Revenue']:
 if col_name in df_sheet.columns:
 idx = df_sheet.columns.get_loc(col_name)
 worksheet.set_column(idx, idx, None, format_currency)
 for col_name in ['Final % of Target', 'Final Occupancy %']:
 if col_name in df_sheet.columns:
 idx = df_sheet.columns.get_loc(col_name)
 worksheet.set_column(idx, idx, None, format_percent)

 elif sheet = = 'Unfixable Conflicts':
 if 'Total Rate' in df_sheet.columns:
 idx = df_sheet.columns.get_loc('Total Rate')
 worksheet.set_column(idx, idx, None, writer.book.add_format({'num_format': '#, ##0.00'}))

 df_changes_final.to_excel(writer, sheet_name = 'Recommended Changes', index = False)
 room_stats_df.to_excel(writer, sheet_name = 'Room Revenue & Occupancy', index = False)

 # Autofit columns for 'Recommended Changes'
 worksheet1 = writer.sheets['Recommended Changes']
 for i, col in enumerate(df_changes_final.columns):
 max_len = max(df_changes_final[col].astype(str).map(len).max(), len(str(col)))
 worksheet1.set_column(i, i, max_len + 2)

 # Autofit columns for 'Room Revenue & Occupancy'
 worksheet2 = writer.sheets['Room Revenue & Occupancy']
 for i, col in enumerate(room_stats_df.columns):
 max_len = max(room_stats_df[col].astype(str).map(len).max(), len(str(col)))
 worksheet2.set_column(i, i, max_len + 2)

 if unfixable_conflicts:
 pd.DataFrame(unfixable_conflicts).to_excel(writer, sheet_name = 'Unfixable Conflicts', index = False)

 print(f"\n✅ Exported results to: {os.path.join(OUTPUT_FOLDER, OUTPUT_FILENAME)}")

 # = = = Add Availability Charts to Excel Sheet = = =
 from openpyxl import load_workbook
 from openpyxl.drawing.image import Image as XLImage
 import glob
 import time

 try:
 full_path = os.path.join(OUTPUT_FOLDER, OUTPUT_FILENAME)

 # Wait briefly to ensure Excel file is fully written
 if not os.path.exists(full_path):
 raise FileNotFoundError(f"Excel output not found at {full_path}")

 time.sleep(2)# Give system time to release file handles
 wb = load_workbook(full_path)

 chart_sheet = wb.create_sheet("Availability Charts")

 row = 1
 chart_paths = sorted(glob.glob(os.path.join(OUTPUT_FOLDER, "Availability_Comparison_*.png")))

 if not chart_paths:
 print("⚠️ No availability charts found to insert into Excel.")

 for chart_path in chart_paths:
 if os.path.exists(chart_path):
 img = XLImage(chart_path)
 img.anchor = f"A{row}"
 chart_sheet.add_image(img)
 row + = 35
 else:
 print(f"⚠️ Missing chart image: {chart_path}")

 final_output_path = full_path.replace(".xlsx", " - Charts.xlsx")
 wb.save(final_output_path)
 print(f"✅ Charts successfully embedded in Excel: {final_output_path}")

 except Exception as e:
 print(f"⚠️ Error embedding charts in Excel: {e}")

 from openpyxl import load_workbook
 from openpyxl.drawing.image import Image as XLImage

 try:
 full_path = os.path.join(OUTPUT_FOLDER, OUTPUT_FILENAME)
 wb = load_workbook(full_path)

 chart_sheet = wb.create_sheet("Availability Charts")

 import glob
 row = 1
 chart_paths = glob.glob(os.path.join(OUTPUT_FOLDER, "Availability_Comparison_*.png"))

 for chart_path in sorted(chart_paths):
 img = XLImage(chart_path)
 img.anchor = f"A{row}"
 chart_sheet.add_image(img)
 row + = 35 # Space between images

 final_output_path = full_path.replace(".xlsx", " - Charts.xlsx")
 wb.save(final_output_path)
 print(f"✅ Charts added to Excel: {final_output_path}")
 except Exception as e:
 print(f"⚠️ Error embedding charts in Excel: {e}")

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

try:
 wb.save(final_output_path)
 print(f"✅ Charts successfully embedded in Excel: {final_output_path}")
except Exception as e:
 print(f"⚠️ Error embedding charts in Excel: {e}")
